# -*- coding: utf-8 -*-
import json
import urllib
import math
import openpyxl
x_pi = 3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 扁率
import requests
import json

class Geocoding:
    def __init__(self, api_key):
        self.api_key = api_key

    def geocode(self, address):
        """
        利用高德geocoding服务解析地址获取位置坐标
        :param address:需要解析的地址
        :return:
        """
        geocoding = {'s': 'rsv3',
                     'key': self.api_key,
                     'city': '全国',
                     'address': address}
        geocoding = urllib.urlencode(geocoding)
        ret = urllib.urlopen("%s?%s" % ("http://restapi.amap.com/v3/geocode/geo", geocoding))
        if ret.getcode() == 200:
            res = ret.read()
            json_obj = json.loads(res)
            if json_obj['status'] == '1' and int(json_obj['count']) >= 1:
                geocodes = json_obj['geocodes'][0]
                lng = float(geocodes.get('location').split(',')[0])
                lat = float(geocodes.get('location').split(',')[1])
                return [lng, lat]
            else:
                return None
        else:
            return None


def gcj02_to_bd09(lng, lat):
    """
    火星坐标系(GCJ-02)转百度坐标系(BD-09)
    谷歌、高德——>百度
    :param lng:火星坐标经度
    :param lat:火星坐标纬度
    :return:
    """
    z = math.sqrt(lng * lng + lat * lat) + 0.00002 * math.sin(lat * x_pi)
    theta = math.atan2(lat, lng) + 0.000003 * math.cos(lng * x_pi)
    bd_lng = z * math.cos(theta) + 0.0065
    bd_lat = z * math.sin(theta) + 0.006
    return [bd_lng, bd_lat]


def bd09_to_gcj02(bd_lon, bd_lat):
    """
    百度坐标系(BD-09)转火星坐标系(GCJ-02)
    百度——>谷歌、高德
    :param bd_lat:百度坐标纬度
    :param bd_lon:百度坐标经度
    :return:转换后的坐标列表形式
    """
    x = bd_lon - 0.0065
    y = bd_lat - 0.006
    z = math.sqrt(x * x + y * y) - 0.00002 * math.sin(y * x_pi)
    theta = math.atan2(y, x) - 0.000003 * math.cos(x * x_pi)
    gg_lng = z * math.cos(theta)
    gg_lat = z * math.sin(theta)
    return [gg_lng, gg_lat]


def wgs84_to_gcj02(lng, lat):
    """
    WGS84转GCJ02(火星坐标系)
    :param lng:WGS84坐标系的经度
    :param lat:WGS84坐标系的纬度
    :return:
    """
    if out_of_china(lng, lat):  # 判断是否在国内
        return lng, lat
    dlat = _transformlat(lng - 105.0, lat - 35.0)
    dlng = _transformlng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    mglat = lat + dlat
    mglng = lng + dlng
    return [mglng, mglat]


def gcj02_to_wgs84(lng, lat):
    """
    GCJ02(火星坐标系)转GPS84
    :param lng:火星坐标系的经度
    :param lat:火星坐标系纬度
    :return:
    """
    if out_of_china(lng, lat):
        return lng, lat
    dlat = _transformlat(lng - 105.0, lat - 35.0)
    dlng = _transformlng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    mglat = lat + dlat
    mglng = lng + dlng
    return [lng * 2 - mglng, lat * 2 - mglat]


def bd09_to_wgs84(bd_lon, bd_lat):
    lon, lat = bd09_to_gcj02(bd_lon, bd_lat)
    return gcj02_to_wgs84(lon, lat)


def wgs84_to_bd09(lon, lat):
    lon, lat = wgs84_to_gcj02(lon, lat)
    return gcj02_to_bd09(lon, lat)


def _transformlat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + \
          0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * pi) + 40.0 *
            math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * pi) + 320 *
            math.sin(lat * pi / 30.0)) * 2.0 / 3.0
    return ret


def _transformlng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + \
          0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * pi) + 40.0 *
            math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 *
            math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
    return ret


def out_of_china(lng, lat):
    """
    判断是否在国内，不在国内不做偏移
    :param lng:
    :param lat:
    :return:
    """
    return not (lng > 73.66 and lng < 135.05 and lat > 3.86 and lat < 53.55)


if __name__ == '__main__':
    lng = 128.543
    lat = 37.065
    result1 = gcj02_to_bd09(lng, lat)
    result2 = bd09_to_gcj02(lng, lat)
    result3 = wgs84_to_gcj02(lng, lat)
    result4 = gcj02_to_wgs84(lng, lat)
    result5 = bd09_to_wgs84(lng, lat)
    result6 = wgs84_to_bd09(lng, lat)
    workbook = openpyxl.load_workbook("G:0525.xlsx")
    sheet = workbook.get_sheet_by_name('0525')
    print(sheet)
    print(sheet.title)

    # 激活的工作表
    print(workbook.active)
    #print(sheet.max_row, sheet.max_column)
    for i in range(1,sheet.max_row+1):#sheet.max_row+1
        Otemln=sheet.cell(row=i, column=1).value#第一行为1
        Otemla=sheet.cell(row=i, column=2).value
        #Dtemln=sheet.cell(row=i, column=3).value#第一行为1
       # Dtemla=sheet.cell(row=i, column=4).value
       # print Otemla,Otemln
        print i
        try:
            Ox=bd09_to_wgs84(float(Otemln),float(Otemla))
            print Ox
            #Dx = bd09_to_wgs84(float(Dtemln), float(Dtemla))
            #Oy= bd09_to_gcj02(float(Otemln),float(Otemla))
            #Dy = bd09_to_gcj02(float(Dtemln), float(Dtemla))
            sheet.cell(row=i, column=5, value=Ox[0])
            sheet.cell(row=i, column=6, value=Ox[1])
            sheet.cell(row=i, column=7, value=Dx[0])
            sheet.cell(row=i, column=8, value=Dx[1])
            '''Ourl = 'http://restapi.amap.com/v3/geocode/regeo?output=json&location='+str(Oy[0])+','+str(Oy[1])+'&key=a0e8201f84716690ae7c07155273d03a&radius=1000&extensions=all'
            #url = 'http://restapi.amap.com/v3/place/polygon?key=a0e8201f84716690ae7c07155273d03a&polygon=' + (coords) + '&keywords='+keywrod+'&types=&offset=20&page='+page+'&extensions=all'
            Or = requests.get(Ourl, auth=('user', 'pass'))
            Or_js = Or.json()
            sheet.cell(row=i, column=9, value=Or_js["regeocode"]["formatted_address"])
            Durl = 'http://restapi.amap.com/v3/geocode/regeo?output=json&location='+str(Dy[0])+','+str(Dy[1])+'&key=a0e8201f84716690ae7c07155273d03a&radius=1000&extensions=all'
            #url = 'http://restapi.amap.com/v3/place/polygon?key=a0e8201f84716690ae7c07155273d03a&polygon=' + (coords) + '&keywords='+keywrod+'&types=&offset=20&page='+page+'&extensions=all'
            Dr = requests.get(Durl, auth=('user', 'pass'))
            Dr_js = Dr.json()
            sheet.cell(row=i, column=10, value=Dr_js["regeocode"]["formatted_address"])'''
        except:
            continue
    #for line in ws:
    workbook.save('G:05252.xlsx')
    #g = Geocoding('a0e8201f84716690ae7c07155273d03a')  # 这里填写你的高德api的key
    #result7 = g.geocode('北京市朝阳区朝阳公园')

